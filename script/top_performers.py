import pandas as pd
import os
import openpyxl # Import openpyxl

month_parcitipated = 4
vependle_threshold = 80000 
top_rank = 100

# Get the absolute path of the directory where the script is located
script_dir = os.path.dirname(os.path.abspath(__file__))
# Define and create output directory
output_dir = os.path.join(script_dir, "output") # Use os.path.join for robust path construction
os.makedirs(output_dir, exist_ok=True) # Create output directory if it doesn't exist

excel_file_path = os.path.join(script_dir, "Pendle_leaderboard.xlsx")

# --- New Data Loading with openpyxl to get hyperlinks ---
workbook = openpyxl.load_workbook(excel_file_path, data_only=True) # data_only=True to get values if there are formulas
xls_data_frames = {}
expected_columns = ["#", "User", "Total Rewards", "ETH per 1000 vePENDLE (APR)", "vePENDLE Balance"] # Define expected columns

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data_for_df = []
    
    # Attempt to find header row and column indices
    header_row = [cell.value for cell in sheet[1]] # Assuming header is in the first row
    column_indices = {}
    try:
        for col_name in expected_columns:
            column_indices[col_name] = header_row.index(col_name)
    except ValueError as e:
        print(f"Warning: Sheet '{sheet_name}' is missing one of the expected columns ({e}). Skipping this sheet.")
        continue # Skip to the next sheet

    # Read data rows
    for row_idx in range(2, sheet.max_row + 1): # Start from the second row
        row_values = {}
        valid_row = True
        for col_name in expected_columns:
            cell_idx = column_indices[col_name] + 1 # openpyxl is 1-indexed for columns
            cell = sheet.cell(row=row_idx, column=cell_idx)
            
            if col_name == "User":
                if cell.hyperlink and cell.hyperlink.target:
                    row_values[col_name] = cell.hyperlink.target
                else:
                    row_values[col_name] = cell.value
            else:
                row_values[col_name] = cell.value
            
            # Basic check if critical data is missing (can be refined)
            if row_values[col_name] is None and col_name != "User": # User can be None if no link and no text
                pass # Allow User to be None for now, will be handled by regex/dropna later
        
        data_for_df.append(row_values)

    if data_for_df: # Only create DataFrame if there's data
        df_from_sheet = pd.DataFrame(data_for_df)
        xls_data_frames[sheet_name] = df_from_sheet

xls = xls_data_frames # This is now our dictionary of DataFrames
# --- End of New Data Loading ---

# 2) For each month, compute earn_rate = total_rewards / vePENDLE_balance * 1000
records = []
for month_name, original_df_from_sheet in xls.items():
    if original_df_from_sheet.empty:
        print(f"Sheet '{month_name}' is empty or was skipped. Skipping processing.")
        continue

    df = original_df_from_sheet.copy()

    # Extract the address from the 'User' column (which should now contain URLs or text)
    df["User"] = df["User"].astype(str) # Ensure string type for regex
    # df["User"] = df["User"].str.lstrip('@') # Likely not needed anymore as we get URL directly
    df["User"] = df["User"].str.extract(r'fadd=([^&]+)', expand=False)
    
    # Ensure only expected columns are present before dropping NA, in case sheets have varying extra columns
    df_processed = df[expected_columns].copy() # Use a copy of the slice
    df_processed.dropna(subset=["User", "Total Rewards", "vePENDLE Balance"], inplace=True)

    if df_processed.empty:
        print(f"No valid data to process in sheet '{month_name}' after address extraction and initial NA drop.")
        continue

    df_cleaned = df_processed.copy()
    
    df_cleaned["Total Rewards"] = df_cleaned["Total Rewards"].astype(str).str.replace(" ETH", "", regex=False)
    df_cleaned["Total Rewards"] = pd.to_numeric(df_cleaned["Total Rewards"], errors='coerce')
    df_cleaned.dropna(subset=["Total Rewards"], inplace=True)

    if df_cleaned.empty:
        print(f"No valid data in sheet '{month_name}' after cleaning 'Total Rewards'.")
        continue
        
    # Ensure vePENDLE Balance is numeric and handle potential non-numeric/NaNs
    df_cleaned["vePENDLE Balance"] = pd.to_numeric(df_cleaned["vePENDLE Balance"], errors='coerce')
    df_cleaned.dropna(subset=["vePENDLE Balance"], inplace=True)
    
    # Filter out zero vePENDLE Balance to prevent DivisionByZeroError
    df_cleaned = df_cleaned[df_cleaned["vePENDLE Balance"] != 0]

    if df_cleaned.empty:
        print(f"No valid data in sheet '{month_name}' after cleaning 'vePENDLE Balance' or all balances are zero.")
        continue

    df_cleaned["earn_rate"] = df_cleaned["Total Rewards"] / df_cleaned["vePENDLE Balance"] * 1_000
    df_cleaned["month"] = month_name
    # Include Total Rewards and vePENDLE Balance for later aggregation
    records.append(df_cleaned[["User", "month", "earn_rate", "Total Rewards", "vePENDLE Balance"]])

# 4) Compute per-user aggregate metrics (this part should be after all_data is created)
# Ensure all_data is defined correctly from the records list before this step
if not records:
    print("No data collected from any sheets. Exiting.")
    exit()
all_data = pd.concat(records, ignore_index=True)

agg = all_data.groupby("User").agg(
    avg_rate=("earn_rate", "mean"),
    med_rate=("earn_rate", "median"),
    best_rate=("earn_rate", "max"),
    months_participated=("earn_rate", "count"),
    sum_total_rewards=("Total Rewards", "sum"),
    avg_total_rewards=("Total Rewards", "mean"),
    med_total_rewards=("Total Rewards", "median"),
    sum_vependle_balance=("vePENDLE Balance", "sum"),
    avg_vependle_balance=("vePENDLE Balance", "mean"),
    med_vependle_balance=("vePENDLE Balance", "median")
).reset_index()

# 5) (Optional) Measure consistency: count how many months each user was in the top-decile of earn_rate
#    first compute percentile within each month (still needed for top_decile_months)
all_data["pct_rank"] = all_data.groupby("month")["earn_rate"].rank(pct=True)
top_decile = all_data[all_data.pct_rank >= 0.90]
consistency = top_decile.groupby("User").size().rename("top_decile_months").reset_index()

# Calculate actual rank number within each month based on earn_rate (higher earn_rate = better/lower rank number)
all_data["rank_in_month"] = all_data.groupby("month")["earn_rate"].rank(method='min', ascending=False)

# Calculate average monthly rank number for each user
avg_monthly_rank_num = all_data.groupby("User")["rank_in_month"].mean().rename("avg_monthly_rank_number").reset_index()
med_monthly_rank_num = all_data.groupby("User")["rank_in_month"].median().rename("median_monthly_rank_number").reset_index()
# Calculate average monthly percentile rank for each user (can be kept if other metrics depend on it, or removed if not)
# avg_monthly_pct_rank = all_data.groupby("User")["pct_rank"].mean().rename("avg_monthly_pct_rank").reset_index()

# 6) Merge all metrics back in
agg = agg.merge(consistency, on="User", how="left")
# agg = agg.merge(avg_monthly_pct_rank, on="User", how="left") # Merge if avg_monthly_pct_rank is still needed elsewhere
agg = agg.merge(avg_monthly_rank_num, on="User", how="left")
agg = agg.merge(med_monthly_rank_num, on="User", how="left")
agg = agg.fillna(0) # Apply fillna once after all merges

# 7) Now you can sort however you like:
#    by highest average earn rate, for those with at least month_parcitipated months:
filtered_agg_for_avg = agg[agg["months_participated"] >= month_parcitipated]
top_by_avg = filtered_agg_for_avg.sort_values("avg_rate", ascending=False).head(top_rank)
top_by_avg.to_csv(os.path.join(output_dir, "top_by_average_earn_rate.csv"), index=False) # Save to output_dir

#    by consistency (months in top 10% based on earn_rate percentile):
top_by_consistency = agg.sort_values("top_decile_months", ascending=False).head(top_rank)
top_by_consistency.to_csv(os.path.join(output_dir, "top_by_consistency.csv"), index=False) # Save to output_dir

#    NEW: by lowest average monthly rank number, for those with at least month_parcitipated months:
filtered_agg_for_rank_num = agg[agg["months_participated"] >= month_parcitipated]
top_by_avg_rank_num = filtered_agg_for_rank_num.sort_values("avg_monthly_rank_number", ascending=True).head(top_rank)
top_by_avg_rank_num.to_csv(os.path.join(output_dir, "top_by_average_monthly_rank.csv"), index=False) # Save to output_dir

# --- New Lists based on High vePENDLE Balance ---
agg_high_vependle = agg[agg["avg_vependle_balance"] > vependle_threshold]

#    (1) Highest average earn rate among those with avg_vependle_balance > threshold
top_high_vependle_by_avg_rate = agg_high_vependle.sort_values("avg_rate", ascending=False).head(top_rank)
top_high_vependle_by_avg_rate.to_csv(os.path.join(output_dir, "top_high_vependle_by_avg_rate.csv"), index=False)

#    (2) Highest average rank among those with avg_vependle_balance > threshold
top_high_vependle_by_avg_rank = agg_high_vependle.sort_values("avg_monthly_rank_number", ascending=True).head(top_rank)
top_high_vependle_by_avg_rank.to_csv(os.path.join(output_dir, "top_high_vependle_by_avg_rank.csv"), index=False)

# Set Pandas display option to show all columns for the printouts
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 200) 

print("--- Top Performers --- (Saved to CSVs in 'output' directory)\n")
print("Top by average earn_rate (min", month_parcitipated, "months participation):\n", top_by_avg)
print("\nTop by consistency (months in top 10% by earn_rate percentile):\n", top_by_consistency)
print("\nTop by average monthly rank number (min", month_parcitipated, "months participation, lower rank is better):\n", top_by_avg_rank_num)
print("\n--- Top Performers with Avg vePENDLE Balance >", vependle_threshold, "---\n")
print("Top by average earn_rate (Avg vePENDLE >", vependle_threshold, "):\n", top_high_vependle_by_avg_rate)
print("\nTop by average monthly rank number (Avg vePENDLE >", vependle_threshold, ", lower rank is better):\n", top_high_vependle_by_avg_rank)